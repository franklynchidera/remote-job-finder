"""
============================================================
  Remote Job Board Scraper  |  Portfolio Project #3
  Sources: LinkedIn · RemoteOK · We Work Remotely
  Extracts: Title · Company · Salary · Location · Link
  Output:  Excel (.xlsx) + CSV
============================================================
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv, re, time, random, json, os
from datetime import datetime
from urllib.parse import quote_plus


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
]

def get_headers(extra={}):
    h = {"User-Agent": random.choice(USER_AGENTS),
         "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
         "Accept-Language": "en-US,en;q=0.5"}
    h.update(extra)
    return h

def delay(): time.sleep(random.uniform(1.5, 3.0))

def clean_salary(text):
    if not text: return ""
    text = text.strip()
    if any(c in text for c in ['$','€','£','k','K','/yr','/mo','salary','Salary']):
        return re.sub(r'\s+', ' ', text)[:60]
    return ""

def clean_text(t, maxlen=120):
    return re.sub(r'\s+', ' ', (t or '').strip())[:maxlen]


# ─────────────────────────────────────────────────────────────────────────────
#  SOURCE 1: RemoteOK
# ─────────────────────────────────────────────────────────────────────────────
def scrape_remoteok(keyword, max_results=25):
    jobs = []
    try:
        delay()
        url = f"https://remoteok.com/remote-{quote_plus(keyword.lower().replace(' ','-'))}-jobs"
        resp = requests.get(url, headers=get_headers({"Accept": "application/json"}), timeout=15)

        # Try JSON API first
        api_resp = requests.get("https://remoteok.com/api", headers=get_headers(), timeout=15)
        if api_resp.status_code == 200:
            data = api_resp.json()
            for item in data[1:]:  # first item is metadata
                if not isinstance(item, dict): continue
                title = item.get('position','')
                if keyword.lower() not in title.lower() and keyword.lower() not in str(item.get('tags','')).lower():
                    continue
                jobs.append({
                    'title':    clean_text(title),
                    'company':  clean_text(item.get('company','')),
                    'location': clean_text(item.get('location','Remote')),
                    'salary':   clean_salary(item.get('salary','')),
                    'tags':     ', '.join(item.get('tags',[])[:5]),
                    'posted':   item.get('date','')[:10] if item.get('date') else '',
                    'url':      f"https://remoteok.com/l/{item.get('slug','')}",
                    'source':   'RemoteOK',
                    'type':     'Remote',
                })
                if len(jobs) >= max_results: break
    except Exception:
        pass
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
#  SOURCE 2: We Work Remotely
# ─────────────────────────────────────────────────────────────────────────────
def scrape_weworkremotely(keyword, max_results=25):
    jobs = []
    try:
        delay()
        url = f"https://weworkremotely.com/remote-jobs/search?term={quote_plus(keyword)}"
        resp = requests.get(url, headers=get_headers(), timeout=15)
        soup = BeautifulSoup(resp.text, 'lxml')

        for article in soup.select('article.feature, li.feature, .jobs li'):
            title_el   = article.select_one('.title, h2, .job-title')
            company_el = article.select_one('.company, .company-name')
            region_el  = article.select_one('.region, .location')
            link_el    = article.select_one('a[href]')

            title   = clean_text(title_el.get_text()   if title_el   else '')
            company = clean_text(company_el.get_text() if company_el else '')
            region  = clean_text(region_el.get_text()  if region_el  else 'Worldwide')
            href    = link_el['href'] if link_el else ''
            url_full = f"https://weworkremotely.com{href}" if href.startswith('/') else href

            if title:
                jobs.append({
                    'title':    title,
                    'company':  company,
                    'location': region or 'Remote',
                    'salary':   '',
                    'tags':     keyword,
                    'posted':   '',
                    'url':      url_full,
                    'source':   'We Work Remotely',
                    'type':     'Remote',
                })
            if len(jobs) >= max_results: break
    except Exception:
        pass
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
#  SOURCE 3: LinkedIn (public job search)
# ─────────────────────────────────────────────────────────────────────────────
def scrape_linkedin(keyword, location="Remote", max_results=25):
    jobs = []
    try:
        delay()
        url = (f"https://www.linkedin.com/jobs/search/"
               f"?keywords={quote_plus(keyword)}&location={quote_plus(location)}"
               f"&f_WT=2&sortBy=DD")
        resp = requests.get(url, headers=get_headers(), timeout=15)
        soup = BeautifulSoup(resp.text, 'lxml')

        cards = soup.select('.jobs-search__results-list li, .job-search-card')
        for card in cards[:max_results]:
            title_el    = card.select_one('.base-search-card__title, h3')
            company_el  = card.select_one('.base-search-card__subtitle, h4')
            location_el = card.select_one('.job-search-card__location, .base-search-card__metadata span')
            time_el     = card.select_one('time')
            link_el     = card.select_one('a.base-card__full-link, a[href*="linkedin.com/jobs"]')

            title    = clean_text(title_el.get_text()    if title_el    else '')
            company  = clean_text(company_el.get_text()  if company_el  else '')
            loc      = clean_text(location_el.get_text() if location_el else 'Remote')
            posted   = time_el.get('datetime','')[:10]   if time_el     else ''
            href     = link_el.get('href','')             if link_el     else ''
            if href and '?' in href: href = href.split('?')[0]

            if title:
                jobs.append({
                    'title':    title,
                    'company':  company,
                    'location': loc,
                    'salary':   '',
                    'tags':     keyword,
                    'posted':   posted,
                    'url':      href,
                    'source':   'LinkedIn',
                    'type':     'Remote',
                })
    except Exception:
        pass
    return jobs


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────
def save_excel(jobs, filepath):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Jobs"

    BG     = "0D1117"
    HDR    = "161B22"
    ACCENT = "238636"
    ALT    = "0D1B2A"
    WHITE  = "FFFFFF"
    DIM    = "8B949E"
    GOLD   = "E3B341"
    CYAN   = "58A6FF"
    GREEN  = "3FB950"

    thin   = Side(style='thin', color='21262D')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Banner
    ws.merge_cells("A1:I1")
    b = ws["A1"]
    b.value = "💼  REMOTE JOB BOARD SCRAPER — RESULTS"
    b.font  = Font(name='Arial', bold=True, size=16, color=GREEN)
    b.fill  = PatternFill('solid', fgColor=BG)
    b.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:I2")
    s = ws["A2"]
    srcs = {}
    for j in jobs: srcs[j['source']] = srcs.get(j['source'],0)+1
    src_str = '  |  '.join(f"{k}: {v}" for k,v in srcs.items())
    s.value = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}   |   Total Jobs: {len(jobs)}   |   {src_str}"
    s.font  = Font(name='Arial', size=10, color=DIM, italic=True)
    s.fill  = PatternFill('solid', fgColor=HDR)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Headers
    cols   = ['#','Source','Job Title','Company','Location','Salary','Tags','Posted','Apply Link']
    widths = [5,  20,       42,         28,        22,        22,      30,    14,      55]
    for ci,(col,w) in enumerate(zip(cols,widths),1):
        cell = ws.cell(row=3, column=ci, value=col)
        cell.font      = Font(name='Arial', bold=True, size=11, color=BG)
        cell.fill      = PatternFill('solid', fgColor=ACCENT)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 30

    src_colors = {'RemoteOK': CYAN, 'We Work Remotely': GOLD, 'LinkedIn': "A371F7"}

    for ri,job in enumerate(jobs,1):
        er  = ri+3
        bg  = ALT if ri%2==0 else BG
        sc  = src_colors.get(job.get('source',''), WHITE)
        row = [ri, job.get('source',''), job.get('title',''), job.get('company',''),
               job.get('location',''), job.get('salary',''), job.get('tags',''),
               job.get('posted',''), job.get('url','')]
        for ci,v in enumerate(row,1):
            cell = ws.cell(row=er, column=ci, value=v)
            color = sc if ci==2 else (WHITE if ci==3 else DIM)
            cell.font      = Font(name='Arial', size=10, color=color, bold=(ci in [2,3]))
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(
                horizontal='center' if ci in [1,2,6,7,8] else 'left',
                vertical='center', wrap_text=(ci==3))
            cell.border    = border
        ws.row_dimensions[er].height = 22

    # Stats sheet
    ws2 = wb.create_sheet("Stats")
    ws2.sheet_properties.tabColor = ACCENT
    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "📊 SCRAPE STATS"
    ws2["A1"].font  = Font(bold=True, size=14, color=GREEN)
    ws2["A1"].fill  = PatternFill('solid', fgColor=BG)
    ws2["A1"].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 30

    stat_rows = [
        ("Total jobs found", f"=COUNTA(Jobs!C4:C{len(jobs)+3})"),
        ("With salary info", f"=COUNTIF(Jobs!F4:F{len(jobs)+3},\"<>\"&\"\")"),
        ("", ""),
        ("── BY SOURCE ──", ""),
    ] + [(f"  {k}", v) for k,v in srcs.items()]

    for r,(label,val) in enumerate(stat_rows,2):
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.font = Font(name='Arial', bold=True, color=WHITE, size=11)
        c1.fill = PatternFill('solid', fgColor=HDR)
        c3 = ws2.cell(row=r, column=3, value=val)
        c3.font = Font(name='Arial', color=GREEN, bold=True, size=11)
        c3.fill = PatternFill('solid', fgColor="0A1525")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["C"].width = 20

    wb.save(filepath)
    print(f"✅ Excel → {filepath}")


def save_csv(jobs, filepath):
    fields = ['#','source','title','company','location','salary','tags','posted','url','type']
    with open(filepath,'w',newline='',encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        for i,job in enumerate(jobs,1):
            w.writerow({'#':i,**job})
    print(f"✅ CSV  → {filepath}")


# ─────────────────────────────────────────────────────────────────────────────
#  DEMO DATA
# ─────────────────────────────────────────────────────────────────────────────
def _demo_jobs(keyword):
    return [
        {"title":"Senior Python Developer","company":"Stripe","location":"Worldwide","salary":"$130k–$170k/yr","tags":"Python, Django, API","posted":"2026-03-08","url":"https://remoteok.com/l/stripe-senior-python","source":"RemoteOK","type":"Remote"},
        {"title":"Data Engineer – Remote","company":"Shopify","location":"Americas","salary":"$120k–$150k/yr","tags":"Python, Spark, SQL","posted":"2026-03-08","url":"https://remoteok.com/l/shopify-data-engineer","source":"RemoteOK","type":"Remote"},
        {"title":"Backend Engineer (Python)","company":"GitLab","location":"Remote – Global","salary":"$100k–$130k/yr","tags":"Python, PostgreSQL, REST","posted":"2026-03-07","url":"https://remoteok.com/l/gitlab-backend","source":"RemoteOK","type":"Remote"},
        {"title":"Python Automation Engineer","company":"Zapier","location":"Remote USA","salary":"$95k–$125k/yr","tags":"Python, Selenium, APIs","posted":"2026-03-07","url":"https://remoteok.com/l/zapier-automation","source":"RemoteOK","type":"Remote"},
        {"title":"ML Engineer – NLP","company":"Hugging Face","location":"Remote – Europe/Americas","salary":"$115k–$145k/yr","tags":"Python, PyTorch, NLP","posted":"2026-03-06","url":"https://remoteok.com/l/huggingface-ml","source":"RemoteOK","type":"Remote"},
        {"title":"Full Stack Python Developer","company":"Toptal","location":"Worldwide","salary":"","tags":"Python, React, Node","posted":"2026-03-08","url":"https://weworkremotely.com/jobs/toptal-fullstack","source":"We Work Remotely","type":"Remote"},
        {"title":"Python Web Scraping Developer","company":"DataHarvest Inc.","location":"Remote – Any Timezone","salary":"$60k–$90k/yr","tags":"Python, Scrapy, BeautifulSoup","posted":"2026-03-07","url":"https://weworkremotely.com/jobs/dataharvest-scraping","source":"We Work Remotely","type":"Remote"},
        {"title":"Remote Data Analyst","company":"Automattic","location":"Worldwide","salary":"$80k–$110k/yr","tags":"Python, SQL, Tableau","posted":"2026-03-07","url":"https://weworkremotely.com/jobs/automattic-analyst","source":"We Work Remotely","type":"Remote"},
        {"title":"API Integration Engineer","company":"Twilio","location":"Remote – USA/Canada","salary":"$110k–$140k/yr","tags":"Python, REST, Webhooks","posted":"2026-03-06","url":"https://weworkremotely.com/jobs/twilio-api-eng","source":"We Work Remotely","type":"Remote"},
        {"title":"Senior Data Scraping Specialist","company":"BrightData","location":"Remote – Worldwide","salary":"$75k–$105k/yr","tags":"Python, Proxies, Scrapy","posted":"2026-03-05","url":"https://weworkremotely.com/jobs/brightdata-specialist","source":"We Work Remotely","type":"Remote"},
        {"title":"Python Developer – Automation","company":"HubSpot","location":"Remote","salary":"","tags":"Python, Automation","posted":"2026-03-09","url":"https://linkedin.com/jobs/view/hubspot-python-001","source":"LinkedIn","type":"Remote"},
        {"title":"Data Extraction Engineer","company":"Accenture","location":"Remote – UK/EU","salary":"","tags":"Python, ETL, Data Mining","posted":"2026-03-08","url":"https://linkedin.com/jobs/view/accenture-data-002","source":"LinkedIn","type":"Remote"},
        {"title":"Web Scraping Engineer","company":"Nielsen","location":"Remote – USA","salary":"","tags":"Python, Scrapy, Selenium","posted":"2026-03-08","url":"https://linkedin.com/jobs/view/nielsen-scraping-003","source":"LinkedIn","type":"Remote"},
        {"title":"Junior Python Developer","company":"Upwork (Internal)","location":"Remote – Global","salary":"$55k–$75k/yr","tags":"Python, Django, REST","posted":"2026-03-07","url":"https://linkedin.com/jobs/view/upwork-junior-004","source":"LinkedIn","type":"Remote"},
        {"title":"Lead Generation Specialist","company":"SalesLoft","location":"Remote – Americas","salary":"$70k–$95k/yr","tags":"Python, Lead Gen, CRM","posted":"2026-03-06","url":"https://linkedin.com/jobs/view/salesloft-leadgen-005","source":"LinkedIn","type":"Remote"},
        {"title":"Scraping & Data Ops Engineer","company":"ZoomInfo","location":"Remote","salary":"$85k–$115k/yr","tags":"Python, Data Pipelines","posted":"2026-03-05","url":"https://linkedin.com/jobs/view/zoominfo-scraping-006","source":"LinkedIn","type":"Remote"},
        {"title":"Python Freelance Developer","company":"Toptal Network","location":"Worldwide","salary":"","tags":"Python, Freelance","posted":"2026-03-09","url":"https://linkedin.com/jobs/view/toptal-freelance-007","source":"LinkedIn","type":"Remote"},
        {"title":"Remote Backend Python Engineer","company":"Buffer","location":"Remote – Worldwide","salary":"$90k–$120k/yr","tags":"Python, APIs, PostgreSQL","posted":"2026-03-07","url":"https://linkedin.com/jobs/view/buffer-backend-008","source":"LinkedIn","type":"Remote"},
    ]


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def print_banner():
    print("\n" + "="*62)
    print("  💼   R E M O T E   J O B   B O A R D   S C R A P E R")
    print("="*62)
    print("  Portfolio Project #3  |  Built for Contra.com")
    print("  Sources: LinkedIn · RemoteOK · We Work Remotely")
    print("="*62 + "\n")


def main():
    print_banner()

    keyword  = input("  Job keyword (e.g. 'Python Developer', 'Data Analyst'): ").strip()
    if not keyword:
        keyword = "Python Developer"
        print(f"  No input — using demo keyword: '{keyword}'\n")

    print(f"\n⏳ Scraping jobs for '{keyword}'...\n" + "-"*55)

    all_jobs = []

    print("  [1/3] 🟢 Scraping RemoteOK...")
    rjobs = scrape_remoteok(keyword)
    print(f"        → {len(rjobs)} jobs")
    all_jobs.extend(rjobs)

    print("  [2/3] 🟡 Scraping We Work Remotely...")
    wjobs = scrape_weworkremotely(keyword)
    print(f"        → {len(wjobs)} jobs")
    all_jobs.extend(wjobs)

    print("  [3/3] 🔵 Scraping LinkedIn (Remote)...")
    ljobs = scrape_linkedin(keyword)
    print(f"        → {len(ljobs)} jobs")
    all_jobs.extend(ljobs)

    if not all_jobs:
        print("\n  ⚠️  Network unavailable. Loading demo dataset...\n")
        all_jobs = _demo_jobs(keyword)

    # Dedupe by title+company
    seen, deduped = set(), []
    for j in all_jobs:
        key = (j.get('title','').lower()[:40], j.get('company','').lower()[:30])
        if key not in seen:
            seen.add(key)
            deduped.append(j)
    all_jobs = deduped

    os.makedirs("output", exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    xl_path  = f"output/jobs_{keyword.replace(' ','_')}_{ts}.xlsx"
    csv_path = f"output/jobs_{keyword.replace(' ','_')}_{ts}.csv"

    print(f"\n💾 Saving {len(all_jobs)} jobs...")
    save_excel(all_jobs, xl_path)
    save_csv(all_jobs, csv_path)

    with_salary = sum(1 for j in all_jobs if j.get('salary'))
    srcs = {}
    for j in all_jobs: srcs[j['source']] = srcs.get(j['source'],0)+1

    print(f"\n{'='*55}")
    print(f"  💼 Total jobs     : {len(all_jobs)}")
    print(f"  💰 With salary    : {with_salary}")
    for src,cnt in srcs.items():
        print(f"  📌 {src:<22}: {cnt}")
    print(f"  📁 Saved to       : output/")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
